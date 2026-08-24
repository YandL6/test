"""
知识库服务 (D 模块)

Phase 1: 文档解析 + 本地检索（无需 RAGFlow 即可工作）
Phase 2: 对接 RAGFlow 实现向量检索
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from loguru import logger

from api_gateway.config import config


class DocumentParser:
    """文档解析器 — 从需求文档中提取文本内容。"""

    @staticmethod
    def parse(filepath: str) -> str:
        """根据文件类型解析文档内容。"""
        path = Path(filepath)
        if not path.exists():
            return ""

        ext = path.suffix.lower()
        try:
            if ext == ".txt" or ext == ".md":
                return path.read_text(encoding="utf-8")
            elif ext == ".json":
                data = json.loads(path.read_text(encoding="utf-8"))
                return json.dumps(data, ensure_ascii=False, indent=2)
            elif ext == ".pdf":
                return DocumentParser._parse_pdf(path)
            elif ext in (".docx", ".doc"):
                return DocumentParser._parse_docx(path)
            elif ext in (".xlsx", ".xls"):
                return DocumentParser._parse_xlsx(path)
            else:
                return path.read_text(encoding="utf-8", errors="ignore")
        except Exception as e:
            logger.warning(f"文档解析失败 {filepath}: {e}")
            return path.read_text(encoding="utf-8", errors="ignore")

    @staticmethod
    def _parse_pdf(path: Path) -> str:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(path))
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            return f"[PDF 文档: {path.name}] - 需安装 PyMuPDF 解析"

    @staticmethod
    def _parse_docx(path: Path) -> str:
        try:
            import docx2txt
            return docx2txt.process(str(path))
        except ImportError:
            return f"[Word 文档: {path.name}] - 需安装 docx2txt 解析"

    @staticmethod
    def _parse_xlsx(path: Path) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True)
            texts = []
            for ws in wb:
                for row in ws.iter_rows(values_only=True):
                    texts.append("\t".join(str(c) if c else "" for c in row))
            wb.close()
            return "\n".join(texts)
        except ImportError:
            return f"[Excel 文档: {path.name}] - 需安装 openpyxl 解析"


class FunctionPointExtractor:
    """从需求文档文本中提取功能点。"""

    # VCU 功能关键词
    KEYWORDS = {
        "档位": ["P挡", "R挡", "N挡", "D挡", "换挡", "DriveReady", "挡位", "组合开关"],
        "扭矩": ["蠕行", "能量回收", "CRBS", "扭矩限制", "减速缓行", "滑行"],
        "上下电": ["上电", "下电", "READY", "启动", "休眠", "唤醒"],
        "安全": ["故障", "报警", "保护", "互锁", "诊断"],
    }

    @staticmethod
    def extract(text: str, domain: str = "general") -> list[dict]:
        """从文本中提取功能点。"""
        points = []
        lines = text.split("\n")
        pid = 1

        keywords = FunctionPointExtractor.KEYWORDS
        if domain != "general" and domain in keywords:
            keywords = {domain: keywords[domain]}

        for line in lines:
            line = line.strip()
            if len(line) < 5:
                continue
            for cat, words in keywords.items():
                for kw in words:
                    if kw in line:
                        points.append({
                            "id": f"FP-{pid:03d}",
                            "category": cat,
                            "keyword": kw,
                            "description": line[:200],
                        })
                        pid += 1
                        break
                else:
                    continue
                break

        # 去重
        seen = set()
        unique = []
        for p in points:
            if p["description"] not in seen:
                seen.add(p["description"])
                unique.append(p)

        return unique


class KnowledgeBaseService:
    """知识库服务 — Phase 1 本地版。"""

    def __init__(self):
        self._documents: dict[str, dict] = {}  # doc_id -> {content, meta}

    def index_document(self, doc_id: str, filepath: str, filename: str) -> dict:
        """索引文档（解析并存入内存）。"""
        content = DocumentParser.parse(filepath)
        meta = {"doc_id": doc_id, "filename": filename, "char_count": len(content)}
        self._documents[doc_id] = {"content": content, "meta": meta}
        logger.info(f"知识库索引文档: {filename} ({len(content)} 字符)")
        return meta

    def search(self, query: str, top_k: int = 5, domain: str | None = None) -> list[dict]:
        """
        本地关键词检索（Phase 1）。

        Phase 2 替换为 RAGFlow 向量检索。
        """
        results = []
        query_lower = query.lower()

        for doc_id, doc in self._documents.items():
            content = doc["content"]
            # 简单关键词匹配 + 上下文提取
            idx = content.lower().find(query_lower)
            if idx >= 0:
                start = max(0, idx - 50)
                end = min(len(content), idx + len(query) + 200)
                snippet = content[start:end].replace("\n", " ").strip()
                score = content.lower().count(query_lower)
                results.append({
                    "doc_id": doc_id,
                    "filename": doc["meta"]["filename"],
                    "snippet": snippet,
                    "score": score,
                })

        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:top_k]

    def get_skill_path(self, domain: str) -> str | None:
        """获取指定域的 SKILL.md 路径。"""
        skill_map = {
            "vcu_gear": "skills/vcu_gear/SKILL.md",
            "vcu_torque": "skills/vcu_torque/SKILL.md",
        }
        rel = skill_map.get(domain)
        if rel:
            full = config.PROJECT_ROOT / rel
            if full.exists():
                return str(full)
        return None


kb_service = KnowledgeBaseService()
